# 读取excel模板文件——分析说明行和表头行，找到需要插入的首行——插入数据——保存新的excel文件


import time
from dataclasses import dataclass
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class FieldSpec:
    """描述 Excel 单列的填充规则（常量或按行号自增）。"""

    increment: bool
    value: str
    number_fmt: str = "_{n}"


def format_cell(spec: FieldSpec, row_index: int) -> str:
    """
    根据列规则与行号（1-based）生成单元格字符串。
    increment 为 True 时在 value 后拼接 number_fmt.format(n=row_index)。
    """
    if not spec.increment:
        return spec.value
    return spec.value + spec.number_fmt.format(n=row_index)


def build_row_values(specs: Sequence[FieldSpec], row_index: int) -> list[str]:
    """按列规则列表生成一行各列字符串。"""
    return [format_cell(spec, row_index) for spec in specs]


def find_data_last_row(sheet: Worksheet) -> int:
    """
    标准函数说明：
    输入：sheet: Worksheet
    输出：int
    功能：找到数据区最后一行，返回行号
    示例：
    >>> sheet = load_workbook("test.xlsx").active
    >>> find_data_last_row(sheet)
    100
    """
    max_row = sheet.max_row or 1
    max_column = sheet.max_column or 1
    for row in range(max_row, 0, -1):
        if any(
            sheet.cell(row=row, column=c).value not in (None, "")
            for c in range(1, max_column + 1)
        ):
            return row
    return 0


def write_data_to_excel(
    file_path: str,
    sheet: Worksheet,
    data_first_row: int,
    data_number: int,
    column_specs: Sequence[FieldSpec],
) -> None:
    """批量写入多行数据；每列取值由 column_specs 决定。"""
    if not column_specs:
        raise ValueError("column_specs 不能为空")
    for i in range(data_number):
        values = build_row_values(column_specs, i + 1)
        row = data_first_row + i
        for col, val in enumerate(values, start=1):
            sheet.cell(row=row, column=col, value=val)


def main():
    # ======= 配置参数 =======
    DEFAULT_COLUMN_SPECS: list[FieldSpec] = [
        FieldSpec(True, "一级分类"),
        FieldSpec(False, "二级分类"),
        FieldSpec(True, "多分类批量名称"),  # true 代表按行号自增
        FieldSpec(True, "简介"),
        FieldSpec(False, "详情链接"),
    ]
    file_path = "C:/Users/YAO/Desktop/python_code/excel-data/data/创建科研辅助工具列表模板 (5).xlsx"
    data_number = 1000
    # ======= 配置参数 =======

    # 打开文件
    workbook = load_workbook(file_path)
    sheet = workbook.active
    # 定位数据区最后一行
    data_last_row = find_data_last_row(sheet)
    # 定位数据区第一行
    data_first_row = data_last_row + 1

    # 写入数据
    write_data_to_excel(
        file_path, sheet, data_first_row, data_number, DEFAULT_COLUMN_SPECS
    )
    # 保存文件，显示年月日时分秒格式的时间戳
    outpath = file_path.replace(
        ".xlsx", f"_{time.strftime('%Y%m%d%H%M%S')}_{data_number}.xlsx"
    )
    workbook.save(outpath)
    # 关闭文件
    workbook.close()
    print(f"Excel文件创建完成: {outpath}")


if __name__ == "__main__":
    main()
